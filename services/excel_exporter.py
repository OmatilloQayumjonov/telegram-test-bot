import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path
from datetime import datetime
import re


def export_results_to_excel(results: list, test_title: str = None, output_filename: str = None) -> str:
    """
    Test natijalarini chiroyli formatlangan professional Excel (.xlsx) fayliga eksport qiladi.
    Alohida test yoki umumiy testlar jamlanmasini qo'llab-quvvatlaydi.
    """
    wb = openpyxl.Workbook()
    ws = wb.active

    # Varaq nomi (Excel cheklovi: max 31 belgi, maxsus belgilarsiz)
    clean_sheet_name = re.sub(r'[\\/*?:\[\]]', '', test_title or "Natijalar").strip()[:30]
    ws.title = clean_sheet_name if clean_sheet_name else "Natijalar"

    # Uslublar (Styles)
    navy_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    gold_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    silver_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    bronze_fill = PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid")

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=13, bold=True, color="1E3A8A")
    subtitle_font = Font(name="Calibri", size=10, italic=True, color="4B5563")

    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )

    if test_title:
        # Alohida test uchun sarlavha paneli
        headers = [
            "№",
            "Talaba Ism-Familiyasi",
            "Telegram Profili",
            "To'plangan Ball",
            "Jami Savollar",
            "Foiz (%)",
            "Baho / Daraja",
            "Topshirilgan Vaqt"
        ]

        ws.cell(row=1, column=1, value=f"📊 TEST NATIJALARI: {test_title.upper()}").font = title_font
        ws.row_dimensions[1].height = 24

        curr_time_str = datetime.now().strftime("%d.%m.%Y %H:%M")
        ws.cell(row=2, column=1, value=f"👥 Jami topshirganlar: {len(results)} nafar | ⏱ Hisobot vaqti: {curr_time_str}").font = subtitle_font
        ws.row_dimensions[2].height = 18

        start_row = 4
        # Jadval bosh sarlavhalari
        for col_num, h in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col_num, value=h)
            cell.fill = navy_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
        ws.row_dimensions[start_row].height = 26

        # Ma'lumotlarni yozish
        row_idx = start_row + 1
        for i, r in enumerate(results, start=1):
            score = r.get("score", 0)
            total = r.get("total", 1) or 1
            percent = round((score / total) * 100, 1)

            if percent >= 86:
                grade = "A'lo (5)"
            elif percent >= 71:
                grade = "Yaxshi (4)"
            elif percent >= 56:
                grade = "Qoniqarli (3)"
            else:
                grade = "Qoniqarsiz (2)"

            medal = "🥇 " if i == 1 else "🥈 " if i == 2 else "🥉 " if i == 3 else ""
            student_name = r.get("full_name") or "Noma'lum"
            full_name = f"{medal}{student_name}"
            username = f"@{r['username']}" if r.get("username") else f"ID: {r.get('user_id', '-')}"

            row_data = [
                i,
                full_name,
                username,
                score,
                total,
                f"{percent}%",
                grade,
                r.get("completed_at", "-")
            ]

            for col_num, val in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_num, value=val)
                cell.border = thin_border
                cell.font = Font(name="Calibri", size=10)

                # Top 3 talabani chiroyli rang bilan belgilash
                if i == 1:
                    cell.fill = gold_fill
                elif i == 2:
                    cell.fill = silver_fill
                elif i == 3:
                    cell.fill = bronze_fill

                if col_num in [1, 4, 5, 6, 7, 8]:
                    cell.alignment = center_align
                else:
                    cell.alignment = left_align

                # Telegram profiliga bosiladigan havola
                if col_num == 3:
                    if r.get("username"):
                        cell.hyperlink = f"https://t.me/{r['username']}"
                        cell.font = Font(name="Calibri", size=10, color="0000FF", underline="single")
                    elif r.get("user_id"):
                        cell.hyperlink = f"tg://user?id={r['user_id']}"
                        cell.font = Font(name="Calibri", size=10, color="0000FF", underline="single")

            ws.row_dimensions[row_idx].height = 20
            row_idx += 1

    else:
        # Barcha testlar jamlanmasi (Umumiy hisobot)
        headers = [
            "№",
            "Talaba Ism-Familiyasi",
            "Telegram Username",
            "Test Nomi",
            "To'plangan Ball",
            "Jami Savollar",
            "Foiz (%)",
            "Baho / Daraja",
            "Topshirilgan Vaqt"
        ]
        ws.append(headers)
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = navy_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
        ws.row_dimensions[1].height = 26

        row_idx = 2
        for i, r in enumerate(results, start=1):
            score = r.get("score", 0)
            total = r.get("total", 1) or 1
            percent = round((score / total) * 100, 1)

            if percent >= 86:
                grade = "A'lo (5)"
            elif percent >= 71:
                grade = "Yaxshi (4)"
            elif percent >= 56:
                grade = "Qoniqarli (3)"
            else:
                grade = "Qoniqarsiz (2)"

            username = f"@{r['username']}" if r.get("username") else f"ID: {r.get('user_id', '-')}"

            row_data = [
                i,
                r.get("full_name", "Noma'lum"),
                username,
                r.get("test_title", "-"),
                score,
                total,
                f"{percent}%",
                grade,
                r.get("completed_at", "-")
            ]
            ws.append(row_data)

            for col_num in range(1, len(row_data) + 1):
                cell = ws.cell(row=row_idx, column=col_num)
                cell.border = thin_border
                if col_num in [1, 3, 5, 6, 7, 8, 9]:
                    cell.alignment = center_align
                else:
                    cell.alignment = left_align
                cell.font = Font(name="Calibri", size=10)

                # Telegram profiliga bosiladigan havola
                if col_num == 3:
                    if r.get("username"):
                        cell.hyperlink = f"https://t.me/{r['username']}"
                        cell.font = Font(name="Calibri", size=10, color="0000FF", underline="single")
                    elif r.get("user_id"):
                        cell.hyperlink = f"tg://user?id={r['user_id']}"
                        cell.font = Font(name="Calibri", size=10, color="0000FF", underline="single")

            row_idx += 1

    # Ustun kengliklarini avtomatik moslashtirish
    for col in ws.columns:
        max_len = 0
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(min(max_len + 4, 35), 12)

    # Fayl nomini belgilash
    if not output_filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        if test_title:
            clean_file_prefix = re.sub(r'[^\w\s-]', '', test_title).strip().replace(' ', '_')
            output_filename = f"data/{clean_file_prefix}_natijalari_{timestamp}.xlsx"
        else:
            output_filename = f"data/barcha_testlar_natijalari_{timestamp}.xlsx"

    out_path = Path(output_filename)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    return str(out_path.resolve())
